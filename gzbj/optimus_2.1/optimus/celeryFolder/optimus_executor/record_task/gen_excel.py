#!/usr/bin/env python
# -*- coding: utf-8 -*-

# @File  : gen_excel.py
# @Author: gaofzhan
# @Email: gaofeng.a.zhang@ericssoin.com
# @Date  : 2021/1/26 14:23
# @Desc  :
import time
import os
import re
import uuid
import base64
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill
from sqlalchemy import and_
# from celery_app import celery_app
from backend.Model.connection import SESSION
from backend.myBluePrint.ericic_v2.model.nova_table import NovaModel
from backend.myBluePrint.ericic_v2.model.cinder_list_table import CinderModel
from backend.myBluePrint.ericic_v2.model.cinder_quota_usage_table import CinderQuotaUsageModel
from backend.myBluePrint.ericic_v2.model.neutron_port_table import NeutronPortTable
from backend.myBluePrint.ericic_v2.model.nova_aggregate_host_relation_table import AggregateHostRelation
from backend.myBluePrint.ericic_v2.model.openstack_project_table import OpenstackProjectModel
from backend.myBluePrint.ericic_v2.model.openstack_quota_table import OpenstackQuotaModel
from backend.myBluePrint.ericic_v2.model.nova_flavor_table import NovaFlavorModel
from backend.myBluePrint.ericic_v2.model.nova_service_table import NovaServiceModel
from backend.myBluePrint.ericic_v2.model.openstack_hypervisor_stats_table import OpenstackHypervisorStatsModel
from backend.myBluePrint.ericic_v2.mongo_post.record_post import RecordPost
from openpyxl.writer.excel import save_virtual_workbook

#  celery_app.send_task('taskModel.method_view.test', task_id=job_id, args=(job_id, cid))
#  base64


# file_name output format 'cid:job_id'
# step1: file_foder  gen excel file
# step2: gened execl file ---> base64  ---> delete file

# step3: call back api base 64 ---> optimus server

base_path = os.path.split(os.path.realpath(__file__))[0]
output_path = base_path + os.sep + '..' + os.sep + '..' + os.sep + "file_folder"


# @celery_app.task
def gen2excel(job_id, cid):
    db_session = SESSION()
    try:
        _flavors = db_session.query(
            NovaFlavorModel
        ).filter(NovaFlavorModel.dc_id == cid).all()
        _vms = db_session.query(
            NovaModel
        ).filter(NovaModel.dc_id == cid).all()
        _hosts = db_session.query(
            NovaServiceModel.host, NovaServiceModel.zone,NovaServiceModel.state
        ).filter(NovaServiceModel.dc_id == cid).all()
        _tenants = db_session.query(
            OpenstackProjectModel
        ).filter(OpenstackProjectModel.dc_id == cid).all()
        _volumes = db_session.query(
            CinderModel
        ).filter(CinderModel.dc_id == cid).all()
        _dc_infos = db_session.query(
            OpenstackHypervisorStatsModel
        ).filter(OpenstackHypervisorStatsModel.dc_id == cid).all()
        _quota_computes = db_session.query(
            OpenstackQuotaModel
        ).filter(OpenstackQuotaModel.dc_id == cid).all()
        _quota_cinders = db_session.query(
            CinderQuotaUsageModel
        ).filter(CinderQuotaUsageModel.dc_id == cid).all()

        def handel_falvors(flavors):
            ret = []
            for flavor in flavors:
                _ = dict()
                _['ID'] = flavor.id
                _['VCPUs'] = flavor.vcpus
                _['RAM'] = flavor.memory_mib
                _['Disk'] = flavor.disk
                ret.append(_)
            return ret

        def handel_vms(vms, db_session):
            ret = []
            for vm in vms:
                _networks = db_session.query(
                    NeutronPortTable.ip_address
                ).filter(vm.id == NeutronPortTable.device_id).all()

                __fliter = '%' + '%s' % vm.host + '%'
                _fliter = and_(
                    AggregateHostRelation.dc_id == vm.dc_id,
                    AggregateHostRelation.host.like(__fliter)
                )
                _az = db_session.query(
                    AggregateHostRelation.availability_zone,
                    AggregateHostRelation.aggregate_name
                ).filter(_fliter).one_or_none()
                _ = dict()
                _['id'] = vm.id
                _['name'] = vm.name
                _['status'] = None
                _['power_state'] = vm.power_state
                _['flavor'] = vm.flavor
                _['created'] = vm.created
                _['host'] = vm.host
                _['az'] = _az[0] if _az else None
                _['ha'] = _az[1] if _az else None
                _['instance_name'] = vm.instance_name
                _['tenant_id'] = vm.tenant_id
                _['networks'] = ','.join([','.join(i) for i in _networks if i])
                ret.append(_)

            return ret

        def handel_hosts(hosts):
            ret = []
            for host in hosts:
                _ = dict()
                _['Hypervisor Hostname'] = host[0]
                _['az'] = host[1]
                _['State'] = host[2]
                ret.append(_)
            return ret

        def handel_tenants(tenants):
            ret = []
            for tenant in tenants:
                _ = dict()
                _['ID'] = tenant.id
                _['Name'] = tenant.name
                ret.append(_)
            return ret

        def handel_volumes(volumes):
            ret = []
            for volume in volumes:
                _ = dict()
                _['ID'] = volume.id
                _['Name'] = volume.name
                _['Status'] = volume.status
                _['Size'] = volume.size
                _['Type'] = volume.volume_type
                _['Bootable'] = volume.bootable
                _['Attached to'] = volume.attach_to
                _['Tenant'] = volume.tenant_id
                ret.append(_)
            return ret

        def handel_dc(dc_infos):
            ret = []
            for dc_info in dc_infos:
                _ = dict()
                _['id'] = dc_info.id
                _['local_gb'] = dc_info.local_gb
                _['local_gb_used'] = dc_info.local_gb_used
                _['memory_mb'] = dc_info.memory_mg
                _['memory_mb_used'] = dc_info.memory_mg_used
                _['vcpus'] = dc_info.vcpus
                _['vcpus_used'] = dc_info.vcpus_used

                _['count'] = None
                _['current_workload'] = None
                _['disk_available_least'] = None
                _['free_disk_gb'] = None
                _['free_ram_mb'] = None
                _['running_vms'] = None

                ret.append(_)
            if len(ret) == 1:
                return ret[0]
            return ret

        def _handel_quota_computes(quota_computes):
            ret = []
            for quota_compute in quota_computes:
                _ = dict()
                _['ID'] = quota_compute.id
                _['Ram In_use'] = quota_compute.ram_in_use
                _['Ram Limit'] = quota_compute.ram_limit
                _['Core In_use'] = quota_compute.cores_in_use
                _['Core Limit'] = quota_compute.cores_limit
                _['Project id'] = quota_compute.project_id
                ret.append(_)
            return ret

        def _handel_quota_cinders(quota_cinders):
            ret = []
            for quota_cinder in quota_cinders:
                _ = dict()
                _['ID'] = quota_cinder.id
                _['Gigabytes In_use'] = quota_cinder.in_use
                _['Gigabytes Limit'] = quota_cinder.limit
                _['Project id'] = quota_cinder.project_id
                ret.append(_)
            return ret

        def get_data(_db_session):
            flavors = handel_falvors(_flavors)
            vms = handel_vms(_vms, _db_session)
            computes = handel_hosts(_hosts)
            tenants = handel_tenants(_tenants)
            volume = handel_volumes(_volumes)
            dc_infos = handel_dc(_dc_infos)

            usage_compute_quata = usage_volume_quata = []
            if tenants:
                usage_compute_quata = _handel_quota_computes(_quota_computes)
                usage_volume_quata = _handel_quota_cinders(_quota_cinders)
            return {
                "flavor": flavors,
                "vm_info": vms,
                "aggregate": None,
                "compute_info": computes,
                "tenants": tenants,
                "volume": volume,
                "dc_info": dc_infos,
                "compute_quota": None,
                "volume_quota": None,
                "usage_compute_quata": usage_compute_quata,
                "usage_volume_quata": usage_volume_quata
            }

        data = get_data(db_session)
    finally:
        db_session.close()
    ret = _write2excel(cid, job_id, data, save=False)
    # write to mongo
    rid = uuid.uuid4().hex
    if ret:
        write2mongo(rid, ret, cid)
        # ret.wb.save('ccccc.xlsx')
    return rid


def write2mongo(rid, content, cid):
    obj = RecordPost(id=rid, content=save_virtual_workbook(content.wb), cid=cid, timestamp=int(time.time()))
    obj.save()


def gen_base64_content(file):
    with open(file, 'rb') as f:
        fd = f.read()
    base64_str = base64.b64encode(fd).decode('utf-8')
    return base64_str


def _write2excel(cid, job_id, data, save=True):
    excel_file = "%s_%s.xlsx" % (cid, job_id)
    _output_infos, sheet_info = prepare(excel_file)
    output_infos = gen_all_data(data, _output_infos, sheet_info, excel_file)
    ret = write2excel(
        output_infos, sheet_info, output_path,
        hide=True,
        hide_list=[
            {
                "name": "VM_Static",
                "hide_field": ['L', 'M', 'N', 'O']
            },
            {
                "name": "Tenant_Static",
                "hide_field": ["G"]},
            {
                "name": "DC_Static",
                "hide_field": ['D', 'H']
            }
        ],
        save=save
    )
    return ret


def gen_all_data(data_infos, output_info, sheet_info, excel_file):
    flavors = data_infos.get('flavor')
    vms = data_infos.get('vm_info')
    ha = data_infos.get('aggregate')
    computes = data_infos.get('compute_info')
    volumes = data_infos.get('volume')
    tenants = data_infos.get('tenants')
    dc_info = data_infos.get('dc_info')
    compute_quota = data_infos.get('compute_quota')
    volume_quota = data_infos.get('volume_quota')
    usage_compute_quata = data_infos.get('usage_compute_quata')
    usage_volume_quata = data_infos.get('usage_volume_quata')

    def call_dc_handel(dc_info):
        tmp = []
        vcpus = dc_info.get("vcpus")
        memory_gb = dc_info.get("memory_mb")
        local_gb = dc_info.get("local_gb")
        volume_gb = None
        memory_used = dc_info.get("memory_mb_used")
        vcpu_usage = dc_info.get("vcpus_used")
        local_usage = dc_info.get("local_gb_used")
        volume_usage = None

        tmp.append(
            [
                vcpus, memory_gb, local_gb, volume_gb,vcpu_usage,memory_used,local_usage,volume_usage
            ]
        )
        return tmp

    def call_tenant_handel(tenants, compute_quota, volume_quota_datas,
                usage_compute_quata, usage_volume_quata):
        tmp = []
        for tenant in tenants:
            tenant_id = tenant.get("ID")
            tenant_name = tenant.get("Name")

            cpu_quota = mem_quota = "No Limit"
            if compute_quota:
                for _compute_quota in compute_quota:
                    if _compute_quota.get("Project ID") == tenant_id:
                        cpu_quota = _compute_quota.get("Cores")
                        mem_quota = _compute_quota.get("Ram")
                        break

            volume_quota = "No Limit"
            if volume_quota_datas:
                for volume_quota_data in volume_quota_datas:
                    if volume_quota_data.get("Project ID") == tenant_id:
                        volume_quota = volume_quota_data.get("Gigabytes")
                        break

            cpu_usage_rate = mem_usage_rate = None
            if usage_compute_quata:
                for _usage_compute_quata in usage_compute_quata:
                    if _usage_compute_quata.get("Project id") == tenant_id:
                        Cpu_In_use = _usage_compute_quata.get("Core In_use")
                        Cpu_Limit = _usage_compute_quata.get("Core Limit")
                        Mem_In_use = _usage_compute_quata.get("Ram In_use")
                        Mem_Limit = _usage_compute_quata.get("Ram Limit")
                        if Cpu_Limit != -1:
                            cpu_usage_rate = int(Cpu_In_use)/int(Cpu_Limit)*100
                        else:
                            cpu_usage_rate = -1
                        if Mem_Limit != -1:
                            mem_usage_rate = int(Mem_In_use)/int(Mem_Limit)*100
                        else:
                            mem_usage_rate = -1
                        cpu_usage_rate = Cpu_In_use
                        mem_usage_rate = Mem_In_use
                        break

            volume_usage_rate = None
            if usage_volume_quata:
                for _usage_volume_quata in usage_volume_quata:
                    if _usage_compute_quata.get("Project id") == tenant_id:
                        Volume_In_use = _usage_volume_quata.get("Gigabytes In_use")
                        Volume_Limit = _usage_volume_quata.get("Gigabytes Limit")
                        volume_usage_rate = int(Volume_In_use)/int(Volume_Limit)*100
                        break
            tmp.append(
                [
                    tenant_name, cpu_quota, mem_quota, volume_quota,cpu_usage_rate,mem_usage_rate,volume_usage_rate
                    # '{:.2f}'.format(cpu_usage_rate)+"%",
                    # '{:.2f}'.format(mem_usage_rate)+"%",
                    # '{:.2f}'.format(volume_usage_rate)+"%"
                ]
            )
        return tmp

    def call_vm_handel(tenants, vms, computes, ha, flavors, volumes):
        tmp = []
        if not computes:
            tmp.append(
                [None for i in range(len(sheet_info.get("VM_Static").get("header")))]
            )
            return tmp
        else:
            for compute in computes:
                compute_status = compute.get('State', None)
                host = compute.get('Hypervisor Hostname', None)
                compute_cpu = compute.get('vCPUs', None)
                compute_mem_mb = compute.get('Memory MB', None)
                compute_disk = compute.get('Total Disk', None)
                compute_numa = None

                HA = None
                AZ = 'nova'
                if ha:
                    for agg in ha:
                        if host in agg.get('hosts'):
                            HA = agg.get('name')
                            AZ = agg.get('availability_zone')
                            break

                none_vms_and_compute = [
                    None, None, None, None, None, None, None,
                    host, HA, AZ, compute_status, compute_cpu,
                    compute_mem_mb, compute_disk, compute_numa,
                    None, None, None, None, None, None, None,
                    None,None,None
                ]
                if not vms:
                    if none_vms_and_compute not in tmp:
                        tmp.append(
                            none_vms_and_compute
                        )
                else:
                    for vm in vms:
                        # get tenant_name
                        tenant_id = vm.get('tenant_id')
                        tenant_name = None
                        for _tenant in tenants:
                            if _tenant.get("ID") == tenant_id:
                                tenant_name = _tenant.get("Name")
                                break

                        if not re.search(vm.get('host'), host):
                            if none_vms_and_compute not in tmp:
                                tmp.append(
                                    none_vms_and_compute
                                )
                        else:
                            if none_vms_and_compute in tmp:
                                tmp.remove(none_vms_and_compute)
                            HA = vm.get('ha', None)
                            AZ = vm.get('az', None)
                            vm_id = vm.get('id')
                            vm_name = vm.get('name')
                            vm_status = vm.get('status')
                            create = vm.get('created')
                            network = vm.get('networks')
                            network = network.replace(",","\n")
                            vm_power_state = vm.get('power_state', None)

                            cpu = ram = disk = None
                            for flavor in flavors:
                                if vm.get('flavor') == flavor.get('ID'):
                                    cpu = flavor.get('VCPUs')
                                    ram = flavor.get('RAM')
                                    disk = flavor.get('Disk')
                                    break

                            volume_name = volume_state = volume_size\
                                = volume_type = volume_bootable = ""
                            volume_total=0
                            if volumes:
                                for volume in volumes:
                                    attachs = volume.get("Attached to", None)
                                    if attachs:
                                        if isinstance(attachs, str):
                                            attach_id = attachs
                                            if attach_id == vm_id:
                                                volume_name += volume.get("Name") + "\n"
                                                volume_state += volume.get("Status") + "\n"
                                                volume_size += str(volume.get("Size")) + "\n"
                                                volume_total += volume.get("Size")
                                                volume_type += str(volume.get("Type")) + "\n"
                                                volume_bootable += str(volume.get("Bootable")) + "\n"
                                        elif isinstance(attachs, (list, tuple)):
                                            for attach in attachs:
                                                attach_id = attach.get("server_id", None)
                                                if attach_id == vm_id:
                                                    volume_name += volume.get("Name") + "\n"
                                                    volume_state += volume.get("Status") + "\n"
                                                    volume_size += str(volume.get("Size")) + "\n"
                                                    volume_type += str(volume.get("Type")) + "\n"
                                                    volume_bootable += str(volume.get("Bootable")) + "\n"

                            if AZ==None: 
                                AZ="nova"
                            tmp.append(
                                [
                                    tenant_name, vm_id, vm_name, vm_status,
                                    vm_power_state, create, network, host,
                                    HA, AZ, compute_status, compute_cpu,
                                    compute_mem_mb, compute_disk, compute_numa,
                                    cpu, 
                                    int(cpu/2)if cpu%2==0 else cpu/2,
                                    ram, disk,
                                    volume_name if volume_name != "" else None if volume_name == '\n' else None,
                                    volume_state if volume_state != "" else None if volume_state == '\n' else None,
                                    volume_size if volume_size != "" else None if volume_size == '\n' else None,
                                    volume_type if volume_type != "" else None if volume_type == '\n' else None,
                                    volume_bootable if volume_bootable != "" else None if volume_bootable == '\n' else None,
                                    volume_total if volume_total != 0 else None if volume_total == '\n' else None,
                                ]
                            )

                temp = []
                for d in tmp:
                    temp.append(d[7])
                if temp.count(host) >= 2:
                    if none_vms_and_compute in tmp:
                        tmp.remove(none_vms_and_compute)
        return tmp

    for k, v in sheet_info.items():
        data = get_vim_init_data(
            output_info, excel_file,
            k,
            v.get("header")
        )
        if k == "DC_Static":
            if dc_info:
                _data = call_dc_handel(dc_info)
                data.extend(_data)
        elif k == "Tenant_Static":
            _data = call_tenant_handel(
                tenants, compute_quota, volume_quota,
                usage_compute_quata, usage_volume_quata
            )
            data.extend(_data)
        elif k == "VM_Static":
            _data = call_vm_handel(
                tenants, vms, computes, ha, flavors, volumes
            )
            data.extend(_data)

    return output_info


class InitWorkbook(object):
    def __init__(self, filename):
        self.name = filename
        if not os.path.exists(self.name):
            self.wb = Workbook()
            self.all_sheets = None
            # try:
            #     self.wb.save(filename=self.name)
            # except PermissionError as e:
            #     print(e)
            #     print("if your hava open file %s, "
            #           "please close the file first. "
            #           "And re-try." % self.name)
            #     exit(1)
        else:
            self.wb = load_workbook(self.name)


class BaseWrokbook(object):
    def __init__(self, filename, workbook, hide=False, hide_list=[]):
        self.name = filename
        self.wb = workbook
        self.hide = hide
        self.hide_list = hide_list
        self.all_sheets = self.wb.sheetnames
        self.alignment = Alignment(wrapText=True,horizontal='center', vertical='center')
        self.title_font = PatternFill("solid", fgColor="7EC0EE")
        self.fromat_title_font = PatternFill("solid", fgColor="FFE4E1")

    def init_sheet(self, titles):
        if isinstance(titles, list):
            for title in reversed(titles):
                # print("all_sheets: %s" % self.all_sheets)
                if not (title in self.all_sheets):
                    self.__dict__['self.ws_%s_%s' % (
                        self.name,
                        title
                    )] \
                        = self.wb.create_sheet(title, 0)
                else:
                    self.__dict__['self.ws_%s_%s' % (
                        self.name,
                        title
                    )] \
                        = self.wb[title]
        else:
            print("Excel file %s, sheet name type error %s; "
                  "it must be list"
                  % (self.name, type(titles))
                  )
            exit(1)

    def get_sheet_wb(self, sheet):
        ws = 'self.ws_{}_{}'.format(self.name, sheet)
        return getattr(self, ws)

    def read(self):
        pass

    def write(self, input, sheet, step=0,
              format_infos=None,
              save=True):
        """
        :param input:
        :param sheet:
        :param step:
        :param format_infos:
            [
                {
                    "position": ["A1", "C1"],
                    "information": "xxxx"
                    "data_location": [row_num, col_num] # [1：1]
                }
            ]
        :return:
        """
        try:
            data = list(zip(*input))
            ws = self.get_sheet_wb(sheet)

            # add sheet title color
            for c in range(len(data)):
                ws[str(get_column_letter(c+1)) + str(1)].fill =\
                    self.fromat_title_font
                ws[str(get_column_letter(c+1)) + str(1+step)].fill =\
                    self.title_font

            row_len = len(data[0])
            col_len = len(data)
            ## format sheet header
            if format_infos:
                for format_info in format_infos:
                    pos = format_info.get('position')
                    location = format_info.get('data_location')
                    ws.merge_cells('%s:%s' % (pos[0], pos[-1]))
                    ws.cell(location[0], location[-1]).value = \
                        format_info.get("information")
            for col in range(col_len):
                row_length = 0
                for row in range(row_len):
                    if step:
                        row = row + step
                    col_datas = str(data[col][row-step]).split()
                    for col_data in col_datas:
                        _row_length =  len(col_data)
                        if _row_length > row_length:
                            row_length = _row_length
                    ws.cell(row=row+1, column=col+1, value=str(data[col][row-step])).alignment = self.alignment
                    if re.search(r"\n", str(data[col][row-step])):
                        ws[str(get_column_letter(col+1)) + str(row)]\
                            .alignment = self.alignment
                column_id = get_column_letter(col+1)
                ws.column_dimensions[column_id].width = row_length+5
            if self.hide:
                for _hides in self.hide_list:
                    if _hides['name'] == sheet:
                        for _hide in _hides['hide_field']:
                            ws.column_dimensions[_hide].hidden = True
            if save:
                self.wb.save(self.name)
        except Exception as e:
            print("write data to excel failed!")
            print(e)
            exit(1)


def get_vim_init_data(infos, filename, name, header=None):
    for info in infos.get(filename):
        if info.get('sheet') == name:
            if header:
                info.get('data').append(header)
            return info.get('data')


def prepare(excel_file):
    output_infos = {
        excel_file: []
    }
    sheet_info = {
        "DC_Static":
            {
                "header":
                    [
                        "Total CPU", "Total Mem\n(MB)", "Total Local Disk\n(GB)",
                        "Total Volume (GB)", "Used CPU", "Used Memory\n(MB)",
                        "Used Disk\n(GB)", "Volume Usage Rate"
                    ],
                "format_info": [
                    {
                        "position": ["A1", "D1"],
                        "information": "Total Hypervisor Resource of Current Datacenter",
                        "data_location": [1, 1]
                    },
                    {
                        "position": ["E1", "H1"],
                        "information": "Utilization Rate of Hypervisor Resource",
                        "data_location": [1, 5]
                    }]
            },
        "Tenant_Static":
            {
                "header":
                    [
                        "Tenant Name", "CPU Quota", "Mem Quota (MB)", "Volume Quota\n(GB)",
                        "Used CPU", "Used Memory\n(MB)", "Volume Usage Rate"
                    ],
                "format_info": [
                    {
                        "position": ["A1", "D1"],
                        "information": "Tenant Quotas",
                        "data_location": [1, 1]
                    },
                    {
                        "position": ["E1", "G1"],
                        "information": "Utilization Rate of Tenant Quotas",
                        "data_location": [1, 5]
                    }]
            },
        "VM_Static":
            {
                "header":
                    [
                        "Tenant Name", "UUID", "Name", "Status", "Power State",
                        "Created Time", "Networks", "Host", "Host Aggregate",
                        "Availability Zone", "Compute State", "infra CPU",
                        "infra Mem", "infra Disk", "NUMA Layout", "vcpu","p-cpu",
                        "Mem\n(MB)", "Disk\n(GB)", "volume name", "volume state",
                        "volume size", "volume type", "bootable","Total Volume Size"
                    ],
                "format_info": [
                    {
                        "position": ["B1", "G1"],
                        "information": "VM Fields",
                        "data_location": [1, 2]
                    },
                    {
                        "position": ["H1", "O1"],
                        "information": "Host Fields",
                        "data_location": [1, 8]
                    },
                    {
                        "position": ["P1", "W1"],
                        "information": "Flavor & Resources",
                        "data_location": [1, 16]
                    }]
            }
    }

    for sheet_name in sheet_info.keys():
        output_infos[excel_file].append(
            {
                'sheet': sheet_name,
                'data': []
            }
        )
    return output_infos, sheet_info


def write2excel(output_infos, sheet_info, output_path, hide=False,
                hide_list=[], save=True):
    for file, sheet_infos in output_infos.items():
        file = output_path + os.sep + file
        if os.path.exists(file):
            os.remove(file)
        workbook = InitWorkbook(file).wb
        wb = BaseWrokbook(file, workbook, hide, hide_list)
        _sheet_infos = []
        for s in sheet_infos:
            if s not in _sheet_infos:
                _sheet_infos.append(s)
        for _sheet_info in _sheet_infos:
            _sheet = _sheet_info.get('sheet', None)
            data = _sheet_info.get('data', None)
            if _sheet and len(_sheet) >= 31:
                sheet = _sheet[-30:]
            else:
                sheet = _sheet
            if sheet:
                wb.init_sheet(titles=[sheet])
            if data and sheet:
                for k, v in sheet_info.items():
                    if k == sheet:
                        format_infos = v.get("format_info")
                wb.write(
                    input=data, sheet=sheet, step=1,
                    format_infos=format_infos,
                    save=save
                )
    return wb if not save else None
