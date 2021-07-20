import os
import re
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill


class InitWorkbook(object):
    def __init__(self, filename):
        self.name = filename
        if not os.path.exists(self.name):
            self.wb = Workbook()
            self.all_sheets = None
            try:
                self.wb.save(filename=self.name)
            except PermissionError as e:
                print(e)
                print("if your hava open file %s, "
                      "please close the file first. "
                      "And re-try." % self.name)
                exit(1)
        else:
            self.wb = load_workbook(self.name)


class BaseWrokbook(object):
    def __init__(self, filename, workbook):
        self.name = filename
        self.wb = workbook
        self.all_sheets = self.wb.sheetnames
        self.alignment = Alignment(wrapText=True)
        self.title_font = PatternFill("solid", fgColor="7EC0EE")

    def init_sheet(self, titles):
        if isinstance(titles, list):
            for title in reversed(titles):
                # print("all_sheets: %s" % self.all_sheets)
                if not (title in self.all_sheets):
                    self.__dict__['self.ws_%s_%s' % (
                        self.name,
                        title
                    )] \
                        = self.wb.create_sheet(title, 0)
                else:
                    self.__dict__['self.ws_%s_%s' % (
                        self.name,
                        title
                    )] \
                        = self.wb[title]
        else:
            print("Excel file %s, sheet name type error %s; "
                  "it must be list"
                  % (self.name, type(titles))
                  )
            exit(1)

    def get_sheet_wb(self, sheet):
        ws = 'self.ws_{}_{}'.format(self.name, sheet)
        return getattr(self, ws)

    def read(self):
        pass

    def write(self, input, sheet):
        try:
            data = list(zip(*input))
            # print("write to excel data %s, sheet: %s" % (data, sheet))
            ws = self.get_sheet_wb(sheet)

            # add sheet title color
            for c in range(len(data)):
                ws[str(get_column_letter(c+1)) + str(1)].fill =\
                    self.title_font

            row_len = len(data[0])
            col_len = len(data)
            for col in range(col_len):
                for row in range(row_len):
                    ws.cell(row=row+1, column=col+1, value=str(data[col][row]))
                    if re.search(r"\n", str(data[col][row])):
                        ws[str(get_column_letter(col+1)) + str(row)]\
                            .alignment = self.alignment
            self.wb.save(self.name)
        except Exception as e:
            print("write data to excel failed!")
            print(e)
            exit(1)


if __name__ == '__main__':
    """
    with open('backend/myBluePrint/ericic/fake_data.json') as f:
        demo_data = f.read()
    demo_data = json.loads(demo_data)
    """
    import uuid
    sheet_infos = [
        {
            'sheet': 'vm_info',
            'data': [
                ('VM ID ', 'VM NAME', 'Status', 'Power State', 'CPU', 'RAM', 'DISK', 'Created Time', 'Networks', 'HOST',
                 'HA', 'AZ', 'Compute state'),
                ['17ac8035-2dc9-44ee-a227-84b94288029b', 'test1', 'ACTIVE', 'Running', 4, 4096, 30,
                 '2020-08-14T08:53:35Z', 'test-net1=192.168.1.12', 'compute1.k2.ericsson.se', None, 'nova', 'up'],
                ['17dbe9ad-e8d0-46b7-b922-eb61280f9a47', 'test2', 'ACTIVE', 'Running', 4, 4096, 30,
                 '2020-08-14T10:10:38Z', 'test-net1=192.168.1.14', 'compute1.k2.ericsson.se', None, 'nova', 'up'],
                ['082f165e-f4a6-4d2b-af15-9a1a7c206da1', 'test3', 'ACTIVE', 'Running', 4, 4096, 30,
                 '2020-08-19T02:49:50Z', 'test-net1=10.1.1.12', 'compute1.k2.ericsson.se', None, 'nova', 'up'],
                ['671768c1-02cc-4207-932a-6809ef320e03', 'test4', 'ACTIVE', 'Running', 4, 4096, 30,
                 '2020-08-19T02:57:22Z', 'test-net1=192.168.1.15,10.1.1.13', 'compute1.k2.ericsson.se', None, 'nova',
                 'up']
            ]
        }
    ]
    file = "../excel/%s.xlsx" % str(uuid.uuid4().hex)
    workbook = InitWorkbook(file).wb
    wb = BaseWrokbook(file, workbook)

    for sheet_info in sheet_infos:
        _sheet = sheet_info.get('sheet', None)
        data = sheet_info.get('data', None)
        if _sheet and len(_sheet) >= 31:
            sheet = _sheet[-30:]
        else:
            sheet = _sheet
        if sheet:
            wb.init_sheet(titles=[sheet])
        if data and sheet:
            wb.write(input=data, sheet=sheet)
